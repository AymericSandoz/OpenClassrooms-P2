# import necessary modules
import requests
from bs4 import BeautifulSoup
import openpyxl
import os
import re
from PIL import Image
from io import BytesIO
from unidecode import unidecode



def main():
    construct_workbook()


def construct_workbook():
    # create an empty workbook
    workbook = openpyxl.Workbook()

    #Get categorys links and create a worksheet by category
    categorys_link = get_categorys_links()
    create_and_fill_worksheet_by_category(categorys_link,workbook)

    # Remove the first sheet(empty) from the Excel file and save it
    first_sheet = workbook['Sheet']
    workbook.remove(first_sheet)
    workbook.save("books.xlsx")



def get_categorys_links():
    # specify the URL
    url = "http://books.toscrape.com/"
    # send a GET request to the URL and get the page content
    page = requests.get(url)
    # parse the page content using BeautifulSoup
    soup = BeautifulSoup(page.content, 'html.parser')
    # find the navigation section
    nav = soup.find(class_='nav-list')
    # get all the category links
    categorys_a = nav.find_all('a')
    # remove the first link, which is not a category
    categorys_a.pop(0)
    return categorys_a


def create_and_fill_worksheet_by_category(categorys_a,workbook):
    # loop through categories
    for category in categorys_a:
        category_file=create_empty_sheet(category,workbook)
        soup=get_index_page_content(category)
        max_page=find_category_page_number(soup)
        column_headers = define_column_headers()
        category_file = write_column_headers(column_headers,category_file)
        get_and_save_category_pages_content(max_page,category,category_file)

def create_empty_sheet(category,workbook):
    # create a new worksheet in the workbook for the current category
    category_file = workbook.create_sheet(title=category.string.strip())
    return category_file


def get_index_page_content(category):
    # get the URL of the category page
    url_page_category = category.get("href")
    # send a GET request to the category page URL
    page = requests.get(f"http://books.toscrape.com/{url_page_category}")
    # parse the category page content using BeautifulSoup
    soup = BeautifulSoup(page.content, 'html.parser')
    return soup
    
 
def find_category_page_number(soup):
    # find the pager section
    pager = soup.find('ul', class_='pager')
    # if the pager section exists, get the page information and the number of page
    if soup.find(class_="pager") is not None: 
        page_info = pager.find('li', class_='current').text.strip()
        max_page = int(page_info.split()[-1])
    # if the pager section doesn't exist, set the number of page to 1
    else: 
        max_page = 1
    
    return max_page

def define_column_headers():
    # create column headers for the worksheet
    coloumn_headers = ["Product page url", "Universal product code", "Title", "Price including tax", "Price excluding tax", "Number available", "Description", "Category", "Review rating", "Image Url"]
    return coloumn_headers

def write_column_headers(coloumn_headers,category_file):   
    # write the column headers to the first row of the worksheet
    for col, header in enumerate(coloumn_headers, start=1):
        category_file.cell(row=1, column=col, value=header) 
    return category_file

def get_and_save_category_pages_content(max_page,category,category_file):
    
    nb_of_articles_by_pages = 20
    # loop through all pages of the current category 
    for page_number in range(1,max_page+1):
        row = 2 + page_number*nb_of_articles_by_pages - nb_of_articles_by_pages
        page_content = get_and_save_category_page_content(max_page,page_number,category,category_file,row)
    


def save_row(coloumn_values,category_file,row):
    # Write the values to the Excel file
    for col, value in enumerate(coloumn_values, start=1):
        category_file.cell(row=row, column=col, value=value)
    return category_file



def get_and_save_category_page_content(max_page,page_number,category,category_file,row):
    # Build the URL of the current category page taking page into account
    
    url_page_category = category.get("href")
    
    if max_page>1:
        url_page_category=url_page_category.replace("index",f"page-{page_number}")
            
        # Send a GET request to the current category page with pagination and create a BeautifulSoup object to parse the HTML content of the page
    page = requests.get(f"http://books.toscrape.com/{url_page_category}")
    soup = BeautifulSoup(page.content, 'html.parser')
    
    get_and_save_category_products_infos(soup,category,category_file,row)
   
    

def get_and_save_category_products_infos(soup,category,category_file,row):
    products = soup.find_all('article')
    # Loop through all the book products on the current page

    
    for product in products:
        get_product_infos_and_save_img(product,category,category_file,row)
        row = row+1

def get_product_infos_and_save_img(product,category,category_file,row):
    # Extract the URL of the current book product 
    products_links=product.find('a')
    product_page_url=products_links.get("href")
    product_page_url = product_page_url.replace("../../../", "http://books.toscrape.com/catalogue/")
        
    # Send a GET request to the current book product page and create a BeautifulSoup object to parse the HTML content of the page
    page = requests.get(product_page_url)
    soup = BeautifulSoup(page.content, 'html.parser')

    # Extract information from the table
    td=soup.find_all('td')
    universal_product_code=td[0].string
    title=soup.find('h1').string
    price_including_tax=convert_price_to_number(td[2].string)
    price_excluding_tax=convert_price_to_number(td[3].string)
    number_available=float(re.search(r'\d+', td[5].string).group()) 
    
    # Find the review rating and extract it
    review_rating = soup.find(class_="star-rating")["class"][1]
    review_rating = convert_review_rating_to_int(review_rating)

    # Find the product description and extract it
    product_description = get_product_description(soup)
    
     # Find the product image and construct the image URL
    image_div=soup.find(id="product_gallery")
    image_url=image_div.find('img')["src"]
    image_url=image_url.replace("../../", "http://books.toscrape.com/")
    save_image(soup,title,category,image_url)


    # Construct a list of values to be written to the Excel file
    coloumn_values=[product_page_url,universal_product_code,title,price_including_tax,price_excluding_tax,number_available,product_description,category.string.strip(),review_rating,image_url]

        # Write the values to the Excel file
    save_row(coloumn_values,category_file,row)
    return coloumn_values


def get_product_description(soup):
    # Find the product description and extract it if it exist
    product_description_div=soup.find(id="product_description")
    if product_description_div is not None:
        product_description = product_description_div.find_next_sibling("p").string
    else:
        product_description = "No product description available"
    return product_description

def save_image(soup,title,category,image_url):
    #construct a name(slud) for later save
    slug_image= generate_slug(title)

    # Create a directory for the images if it does not exist
    repertoire = "images/" + category.string.strip()
    if not os.path.exists(repertoire):
        os.makedirs(repertoire)

    # Download and save the image
    if os.path.isfile(os.path.join(repertoire, slug_image + ".jpg")):
        print(f"Error : Img :{slug_image} already saved")
    else:
        response = requests.get(image_url)
        image = Image.open(BytesIO(response.content))
        image.save(repertoire + "/" + slug_image + ".jpg", "JPEG")


def convert_price_to_number(price):
    float_value = float(price.replace("£", ""))
    return float_value

def convert_review_rating_to_int(review_rating):
    rating_dict = {
    "One": 1,
    "Two": 2,
    "Three": 3,
    "Four": 4,
    "Five": 5
    }


    if review_rating in rating_dict:
        numerical_rating = rating_dict[review_rating]
        return numerical_rating
    else:
        numerical_rating = "Invalid rating"


# define a function to generate slug
def generate_slug(text):
    text = text.lower()
    # remove special characters except hyphen and whitespace
    text = re.sub(r'[^\w\s-]', '', text)
    # replace whitespaces with hyphen
    text = re.sub(r'\s+', '-', text)
    # replace consecutive hyphens with a single hyphen
    text = re.sub(r'--+', '-', text)
    # remove accents
    text = unidecode(text)
    # remove hyphens at the beginning and end of text
    text = text.strip('-')
    # Keep text length < 50
    if len(text) > 50:
        text = text[:47] + "..."
    return text

if __name__ == "__main__": #quel est le nom de l'objet en cours (point d'entrée normal dans un code python)
    main()