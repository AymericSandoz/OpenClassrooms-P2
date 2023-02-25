# import necessary modules
import requests
from bs4 import BeautifulSoup
import csv
import openpyxl
import os
import re
from PIL import Image
from io import BytesIO
import re

# define a function to generate slug
def generate_slug(text):
    text = text.lower()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'\s+', '-', text)
    text = re.sub(r'--+', '-', text)
    text = text.strip('-')
    return text

# create an empty workbook
worbook = openpyxl.Workbook()
books_data = {}

# specify the URL
url = "http://books.toscrape.com/"
# send a GET request to the URL and get the page content
page = requests.get(url)
# parse the page content using BeautifulSoup
soup = BeautifulSoup(page.content, 'html.parser')
# find the navigation section
nav = soup.find(class_='nav-list')
# get all the category links
categorys_a = nav.find_all('a')
# remove the first link, which is not a category
categorys_a.pop(0)


for category in categorys_a[:5]:
 url_page_category = category.get("href")
 page = requests.get(f"http://books.toscrape.com/{url_page_category}")
 soup = BeautifulSoup(page.content, 'html.parser')
 products=soup.find_all('article')
 category_file = worbook.create_sheet(title=category.string.strip())
 
 pager = soup.find('ul', class_='pager')
 if soup.find(class_="pager") is not None: 
  page_info = pager.find('li', class_='current').text.strip()
  max_page = int(page_info.split()[-1])
 else : 
  max_page=1

 coloumn_headers=["product_page_url","universal_product_code","title","price_including_tax","price_excluding_tax","number_available","description","category","review_rating","image_url"]
 for col, header in enumerate(coloumn_headers, start=1):
  category_file.cell(row=1, column=col, value=header) 
 
 row=2
 for page_number in range(1,max_page+1):
  url_page_category = category.get("href")
  print(max_page)
  if max_page>1:
   url_page_category=url_page_category.replace("index",f"page-{page_number}")
  page = requests.get(f"http://books.toscrape.com/{url_page_category}")
  print(url_page_category)
  soup = BeautifulSoup(page.content, 'html.parser')
  products=soup.find_all('article')
  for product in products[:3]:
  
   products_links=product.find('a')
   product_page_url=products_links.get("href")
   product_page_url = product_page_url.replace("../../../", "http://books.toscrape.com/catalogue/")
   page = requests.get(product_page_url)
   soup = BeautifulSoup(page.content, 'html.parser')
   product_information=soup.find(class_='table-striped')

   td=soup.find_all('td')
   universal_product_code=td[0].string
   title=soup.find('h1').string
   price_including_tax=td[2].string
   price_excluding_tax=td[3].string
   number_available=re.search(r'\d+', td[5].string).group()
   product_description_div=soup.find(id="product_description")
   if product_description_div is not None:
    product_description = product_description_div.find_next_sibling("p").string
   else:
    product_description = "No product description available"
   review_rating=soup.find(class_="star-rating")["class"][1]

   image_div=soup.find(id="product_gallery")
   image_url=image_div.find('img')["src"]
   image_url=image_url.replace("../../", "http://books.toscrape.com/")
   slug_image= generate_slug(title)
   
   repertoire = "images/" + category.string.strip()
   if not os.path.exists(repertoire):
    os.makedirs(repertoire)


   if os.path.isfile(os.path.join(repertoire, slug_image + ".jpg")):
    print("Img already saved")
   else:
    response = requests.get(image_url)
    image = Image.open(BytesIO(response.content))
    image.save(repertoire + "/" + slug_image + ".jpg", "JPEG")

   
  
   coloumn_values=[product_page_url,universal_product_code,title,price_including_tax,price_excluding_tax,number_available,product_description,category.string.strip(),review_rating,image_url]
   for col, value in enumerate(coloumn_values, start=1):
    category_file.cell(row=row, column=col, value=value)
   row=row+1



first_sheet = worbook['Sheet']
worbook.remove(first_sheet)
worbook.save("books.xlsx")






